from openpyxl import *
import random
import time

filename = 'N2N1核心词汇.xlsx'

# 读取.xlsx文件
wb = load_workbook(filename)

# 读取Excel里的第一张表
sheet = wb[wb.sheetnames[0]]

# # 遍历B列
# for i in range(1, sheet.max_row+1):
#     print(sheet["B"+str(i)].value)

print('--------------------------------------------------')
print('-----------答题开始，请关闭单词Excel--------------')

epoch_null = False
cols = ['C','D','E']
allAnswers = [i for i in range(2,sheet.max_row+1)]
for epoch in range(10000):
    if epoch%12 == 0 or epoch_null:
        epoch_null = False
        epoch_answers = random.sample(allAnswers, 30) 
    begin_time = time.clock()
    q = random.sample(cols, 2) 

    weights = [max(1,sheet['H'+str(i)].value*2 - sheet['G'+str(i)].value  + 10) for i in epoch_answers]
    r = random.randint(1,sum(weights) + 1)
    s = 0
    row = 0
    for i,j in enumerate(weights):
        s += j
        if s >= r:
            row = epoch_answers[i]
            #print(j,sum(weights)/len(weights))
            break

    wrongAnswers = [i for i in epoch_answers]
    del wrongAnswers[wrongAnswers.index(row)]#删除正确的答案
    wrongAnswers=random.sample(wrongAnswers,3)#随机选取三个错误答案
    answersOptions=wrongAnswers+[row,]#将正确答案和错误答案连接起来
    random.shuffle(answersOptions)#打乱四个答案的顺序
    print('--------------------------------------------------')
    print('问题'+str(epoch+1)+':  '+sheet[q[0]+str(row)].value)
    print('1.'+sheet[q[1]+str(answersOptions[0])].value)
    print('2.'+sheet[q[1]+str(answersOptions[1])].value)
    print('3.'+sheet[q[1]+str(answersOptions[2])].value)
    print('4.'+sheet[q[1]+str(answersOptions[3])].value)
    end_time = time.clock()
    #print('use time ',(end_time-begin_time))
    while 1:
        try:
            ans = int(input('你的选择： '))
            break
        except:
            print('请输入1-4的有效数字~')
            pass
    if(row == answersOptions[ans-1]):
        print('\033[92m答对了！\33[0m')
        sheet['G'+str(row)].value += 1
        wb.save(filename)

    else:
        print('\033[91m答错了!\33[0m')
        sheet['H'+str(row)].value += 1
        wb.save(filename)

    print('['+sheet['A'+str(row)].value+']  '+sheet['C'+str(row)].value+'('+sheet['D'+str(row)].value+')',
          '\033[94m['+sheet['I'+str(row)].value+']  \33[0m',sheet['E'+str(row)].value,
           sheet['G'+str(row)].value,'/',sheet['H'+str(row)].value)
    print(sheet['J'+str(row)].value,' / ',sheet['K'+str(row)].value,)



# 出现频率：  常错 > 未背 > 常对 
