# -*- coding: utf-8 -*-



def mywords():
    from openpyxl import Workbook
    from openpyxl import load_workbook
    import random
    wb = load_workbook('N3wordDAY1.xlsx')
    wb.guess_types = True   #猜测格式类型
    ws=wb.active

    cnt = 0
    truecnt = 0
    falsecnt = 0
    while 1:
        a = random.sample(range(2,ws.max_row+2), 4)
        word = ws[a[0]]
        random.shuffle(a)
        print('------第',cnt+1,'题------')
        print('请问单词\033[0;32;40m',word[0].value,'\033[0m的读音为:')

        for i,v in enumerate(a):
            print(i+1,ws[v][2].value,'  ',end=' ')
        n = 0
        while 1:
            try:
                n = int(input('\n请选择答案编号,退出请按0:'))
                if n<=0 or n >4 :continue
                break
            except:
                print('编号错误')
                continue
        if n == 0: 
            print('------游戏结束------')
            break
        
        #print('题目为: ',word[0].value,'您选择了: ',ws[a[n-1]][2].value)
        print('所有单词为:')
        for i in a:
            print(ws[i][0].value,ws[i][1].value,ws[i][2].value,end='')
        print('')
        cnt += 1
        if word[2] == ws[a[n-1]][2]:
            #print('√√√√√√√√√√√恭喜你~~对了√√√√√√√√√√√')
            truecnt += 1
            print("\033[0;32;40m恭喜你~答对了  共答题",cnt,"道,答对",truecnt,"道\033[0m")  
            #\033[显示方式;前景色;背景色m + 结尾部分：\033[0m
        else:
            #print('×××××××××××不对啊啊啊啊×××××××××××')
            falsecnt += 1
            print("\033[0;31;40m不好意思，答错了  共答题",cnt,"道,答对",truecnt,"道\033[0m")  
        print('')
        print('')

        

if __name__ == '__main__':
    mywords()